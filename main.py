import tkinter as tk
from tkinter import ttk, messagebox
import random
import openpyxl
from openpyxl import Workbook

class BillManagementSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("Bill Management System")
        self.root.geometry("900x600")

        # Variables
        self.customer_name = tk.StringVar()
        self.customer_phone = tk.StringVar()
        self.bill_number = tk.StringVar()
        self.bill_number.set(f"{random.randint(1000, 9999)}")
        self.search_bill = tk.StringVar()
        self.item_name = tk.StringVar()
        self.item_price = tk.DoubleVar()
        self.item_quantity = tk.IntVar()
        self.total_price = tk.StringVar()
        self.tax = tk.StringVar()
        self.final_total = tk.StringVar()

        # Excel file setup
        self.file_name = "bills.xlsx"
        self.setup_excel()

        # Frame for customer details
        customer_frame = ttk.LabelFrame(self.root, text="Customer Details")
        customer_frame.pack(fill="x", padx=10, pady=5)

        ttk.Label(customer_frame, text="Customer Name:").grid(row=0, column=0, padx=5, pady=5)
        ttk.Entry(customer_frame, textvariable=self.customer_name).grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(customer_frame, text="Phone Number:").grid(row=0, column=2, padx=5, pady=5)
        ttk.Entry(customer_frame, textvariable=self.customer_phone).grid(row=0, column=3, padx=5, pady=5)

        ttk.Label(customer_frame, text="Bill Number:").grid(row=0, column=4, padx=5, pady=5)
        ttk.Entry(customer_frame, textvariable=self.search_bill).grid(row=0, column=5, padx=5, pady=5)

        ttk.Button(customer_frame, text="Search", command=self.search_bill_function).grid(row=0, column=6, padx=5, pady=5)

        # Frame for item details
        item_frame = ttk.LabelFrame(self.root, text="Item Details")
        item_frame.pack(fill="x", padx=10, pady=5)

        ttk.Label(item_frame, text="Item Name:").grid(row=0, column=0, padx=5, pady=5)
        ttk.Entry(item_frame, textvariable=self.item_name).grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(item_frame, text="Price:").grid(row=0, column=2, padx=5, pady=5)
        ttk.Entry(item_frame, textvariable=self.item_price).grid(row=0, column=3, padx=5, pady=5)

        ttk.Label(item_frame, text="Quantity:").grid(row=0, column=4, padx=5, pady=5)
        ttk.Entry(item_frame, textvariable=self.item_quantity).grid(row=0, column=5, padx=5, pady=5)

        ttk.Button(item_frame, text="Add Item", command=self.add_item).grid(row=0, column=6, padx=5, pady=5)

        # Frame for bill area
        bill_frame = ttk.Frame(self.root)
        bill_frame.pack(fill="both", expand=True, padx=10, pady=5)

        self.bill_text = tk.Text(bill_frame, wrap="word")
        self.bill_text.pack(fill="both", expand=True)

        # Frame for bill summary and actions
        summary_frame = ttk.LabelFrame(self.root, text="Bill Summary")
        summary_frame.pack(fill="x", padx=10, pady=5)

        ttk.Label(summary_frame, text="Total Price:").grid(row=0, column=0, padx=5, pady=5)
        ttk.Entry(summary_frame, textvariable=self.total_price, state="readonly").grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(summary_frame, text="Tax (5%):").grid(row=0, column=2, padx=5, pady=5)
        ttk.Entry(summary_frame, textvariable=self.tax, state="readonly").grid(row=0, column=3, padx=5, pady=5)

        ttk.Label(summary_frame, text="Final Total:").grid(row=0, column=4, padx=5, pady=5)
        ttk.Entry(summary_frame, textvariable=self.final_total, state="readonly").grid(row=0, column=5, padx=5, pady=5)

        ttk.Button(summary_frame, text="Print Bill", command=self.print_bill).grid(row=0, column=6, padx=5, pady=5)
        ttk.Button(summary_frame, text="Clear", command=self.clear_bill).grid(row=0, column=7, padx=5, pady=5)
        ttk.Button(summary_frame, text="Exit", command=root.quit).grid(row=0, column=8, padx=5, pady=5)

        self.items = []
        self.update_bill_area()

    def setup_excel(self):
        try:
            wb = openpyxl.load_workbook(self.file_name)
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.title = "Bills"
            ws.append(["Bill Number", "Customer Name", "Phone Number", "Items", "Total Price", "Tax", "Final Total"])
            wb.save(self.file_name)

    def add_item(self):
        name = self.item_name.get()
        price = self.item_price.get()
        quantity = self.item_quantity.get()

        if not name or price <= 0 or quantity <= 0:
            messagebox.showerror("Error", "Please fill all item fields with valid data")
            return

        total = price * quantity
        self.items.append((name, price, quantity, total))
        self.update_bill_area()

        self.item_name.set("")
        self.item_price.set(0)
        self.item_quantity.set(0)

    def update_bill_area(self):
        self.bill_text.delete(1.0, tk.END)

        if not self.items:
            self.bill_text.insert(tk.END, "No items added to the bill\n")
            return

        self.bill_text.insert(tk.END, f"Bill Number: {self.bill_number.get()}\n")
        self.bill_text.insert(tk.END, f"Customer Name: {self.customer_name.get()}\n")
        self.bill_text.insert(tk.END, f"Phone Number: {self.customer_phone.get()}\n")
        self.bill_text.insert(tk.END, "\nItems:\n")
        self.bill_text.insert(tk.END, f"{'Item':<20}{'Price':<10}{'Quantity':<10}{'Total':<10}\n")
        self.bill_text.insert(tk.END, f"{'-'*50}\n")

        total_price = 0
        for item in self.items:
            name, price, quantity, total = item
            total_price += total
            self.bill_text.insert(tk.END, f"{name:<20}{price:<10}{quantity:<10}{total:<10}\n")

        tax = total_price * 0.05
        final_total = total_price + tax

        self.total_price.set(f"{total_price:.2f}")
        self.tax.set(f"{tax:.2f}")
        self.final_total.set(f"{final_total:.2f}")

    def print_bill(self):
        if not self.items:
            messagebox.showerror("Error", "No items to print")
            return

        self.save_to_excel()
        print_data = self.bill_text.get(1.0, tk.END)
        print_window = tk.Toplevel(self.root)
        print_window.title("Print Bill")
        print_window.geometry("600x400")

        text_area = tk.Text(print_window, wrap="word")
        text_area.insert(1.0, print_data)
        text_area.pack(fill="both", expand=True)

    def save_to_excel(self):
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb["Bills"]

        items_data = ", ".join([f"{name} ({quantity} x {price})" for name, price, quantity, _ in self.items])
        ws.append([
            self.bill_number.get(),
            self.customer_name.get(),
            self.customer_phone.get(),
            items_data,
            float(self.total_price.get()),
            float(self.tax.get()),
            float(self.final_total.get())
        ])

        wb.save(self.file_name)
        messagebox.showinfo("Saved", "Bill saved to Excel file")

    def clear_bill(self):
        self.customer_name.set("")
        self.customer_phone.set("")
        self.bill_number.set(f"{random.randint(1000, 9999)}")
        self.search_bill.set("")
        self.items.clear()
        self.update_bill_area()

    def search_bill_function(self):
        bill_no = self.search_bill.get()
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb["Bills"]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == bill_no:
                _, customer_name, customer_phone, items, total_price, tax, final_total = row
                self.bill_text.delete(1.0, tk.END)
                self.bill_text.insert(tk.END, f"Bill Number: {bill_no}\n")
                self.bill_text.insert(tk.END, f"Customer Name: {customer_name}\n")
                self.bill_text.insert(tk.END, f"Phone Number: {customer_phone}\n")
                self.bill_text.insert(tk.END, "\nItems:\n")
                self.bill_text.insert(tk.END, items + "\n")
                self.bill_text.insert(tk.END, f"\nTotal Price: {total_price}\n")
                self.bill_text.insert(tk.END, f"Tax: {tax}\n")
                self.bill_text.insert(tk.END, f"Final Total: {final_total}\n")
                return

        messagebox.showerror("Error", "Bill not found")

if __name__ == "__main__":
    root = tk.Tk()
    app = BillManagementSystem(root)
    root.mainloop()

